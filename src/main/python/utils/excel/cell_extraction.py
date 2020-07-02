import re

import pandas as pd
from openpyxl import load_workbook


def convert_to_number(s):
    return int(sum(map(int, str((int.from_bytes(s.encode(), 'little'))))))


file_path = '/Users/navomsaxena/Downloads/ICE_budget_sheet_loading_/' \
            '7000009_Vifor_Diamond_WO_Internal_Budget_V.12_14Dec2018 WBS3.0.xlsm'

file_id = convert_to_number(file_path)

df2 = pd.DataFrame(columns=["fileID", "sheetID", "ColNum", "Col", "RowNum", "Value"])
workbook = load_workbook(file_path, read_only=True, data_only=True)

formats = ['#,##', '409', '%', 'yy', 'General']
i = 0
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    for row in sheet.rows:
        for cell in row:
            if cell.value is not None:
                fmt = cell.number_format
                currency = re.search(r"[€£$]", fmt)
                value = cell.value
                if currency:
                    value = f"{currency.group(0)} {cell.value}"
                if '%' in fmt:
                    value = f"{value * 100} %"
                if not any(f in fmt for f in formats):
                    value = f"{value} {fmt}"
                row_num = cell.row
                column_num = cell.column
                column = cell.column_letter
                df2.loc[i] = [file_id, convert_to_number(sheet_name), column_num, column, row_num, value]
                i = i + 1
